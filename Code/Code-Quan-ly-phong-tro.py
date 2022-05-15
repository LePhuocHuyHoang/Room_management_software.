import tkinter as tk
from tkinter import ttk
from PIL import ImageTk
from tkcalendar import DateEntry
from tkinter import filedialog

import pandas as pd
from openpyxl import *
from tkinter.messagebox import  showinfo
from tkinter import messagebox
import time


window = tk.Tk()
window.geometry('600x350')
window.title('Dang nhap ung dung')
a = ImageTk.PhotoImage(file='D:/duan/anhnen1.jpg')
lb = tk.Label(window, image=a)
lb.place(x=0,y=0)

def login():
    username = entry1.get()
    password = entry2.get()

    if (username == '' and password == ''):
        messagebox.showinfo('That bai!', 'Chua nhap tai khoan va mat khau')
    elif (username == 'quanlytro' and password == '1'):
        messagebox.showinfo('Thanh cong!', 'Chao mung ban')

        window = tk.Tk()
        window.geometry('1020x1000')
        window.title('Pham mem quan ly thong tin khach hang den thue tro')


        def open():
            path = filedialog.askopenfilename()
            df = pd.read_excel(path)
            print(df)

        def save():
            Tenkh = txttkh.get()
            Gt = combogt.get()
            Ns = calNs.get()
            Nn = txtnn.get()
            Dc = txtdc.get()
            Sdt = txtsdt.get()
            Qq = comboqq.get()
            Tdc = txttdc.get()
            Gp = txtgp.get()
            Cccd = txtcccd.get()
            Lhd = combolhd.get()
            Nnp = calnnp.get()
            Ntp = calntp.get()
            Lp = combolp.get()
            Ps = combops.get()
            Tt = combott.get()

            Sdtisnum_true = Sdt.isnumeric()
            Cccdisnum_true = Cccd.isnumeric()
            Tenkhistitle_true = Tenkh.istitle()

            if Sdtisnum_true and Tenkhistitle_true  and Cccdisnum_true:
                wb = Workbook()
                ws = wb.active
                ws['A1'] = 'Ten Khach Hang'
                ws['B1'] = 'Gioi Tinh'
                ws['C1'] = 'Nam Sinh'
                ws['D1'] = 'Que quan'
                ws['E1'] = 'So CCCD'
                ws['F1'] = 'Nghe nghiep'
                ws['G1'] = 'Dia chi'
                ws['H1'] = 'So dien thoai'
                ws['I1'] = 'Loai hop dong'
                ws['J1'] = 'Loai phong'
                ws['K1'] = 'Phong so'
                ws['L1'] = 'So tien dat coc'
                ws['M1'] = 'Gia phong'
                ws['N1'] = 'Ngay nhan phong'
                ws['O1'] = 'Ngay tra phong'
                ws['P1'] = 'Thanh toan'
                ws['A2'] = Tenkh
                ws['B2'] = Gt
                ws['C2'] = Ns
                ws['D2'] = Qq
                ws['E2'] = Cccd
                ws['F2'] = Nn
                ws['G2'] = Dc
                ws['H2'] = Sdt
                ws['I2'] = Lhd
                ws['J2'] = Lp
                ws['K2'] = Ps
                ws['L2'] = Tdc
                ws['M2'] = Gp
                ws['N2'] = Nnp
                ws['O2'] = Ntp
                ws['P2'] = Tt
                wb.save(r'D:\duan\duan1.xlsx')
                showinfo('Thong Bao!', 'Da luu thong tin khach hang')
                file1 = pd.read_excel('duan.xlsx')
                file2 = pd.read_excel('duan1.xlsx')
                all = [file1, file2]
                append = pd.concat(all)
                append.to_excel('duan.xlsx', index=False)
            else:
                showinfo('Thong Bao!','Ban nhap chua dung, hay thu lai!')
        def xnclear():
            xnclear = messagebox.askyesno('Xac nhan!', 'Ban co muon xoa du lieu khong?')
            if xnclear == True:
                    txttkh.delete(0, tk.END)
                    calNs.delete(0, tk.END)
                    comboqq.delete(0, tk.END)
                    txtnn.delete(0, tk.END)
                    txtdc.delete(0, tk.END)
                    txtsdt.delete(0, tk.END)
                    txtcccd.delete(0, tk.END)
                    combolhd.delete(0, tk.END)
                    combogt.delete(0, tk.END)
                    comboqq.delete(0, tk.END)
                    calnnp.delete(0, tk.END)
                    calntp.delete(0, tk.END)
                    txttdc.delete(0, tk.END)
                    txtgp.delete(0, tk.END)
                    combops.delete(0, tk.END)
                    combott.delete(0, tk.END)
                    combolp.delete(0, tk.END)
            elif xnclear==False:
                pass




        def quit():
            res = messagebox.askyesno('Xac nhan!', 'Ban co muon thoat khong?')
            if res == True:
                messagebox.showinfo('Dem', 'Man hinh se tat sau 5 giay')
                time.sleep(5)
                window.quit()

            elif res == False:
                pass


        lblqlkh = tk.Label(window, text='Quan Ly Khach Hang', fg='Blue', font=('Arial bold', 10))
        lblqlkh.grid(column=1, row=1)

        lblhd = tk.Label(window, text='Quan Ly Hop Dong', fg='Blue', font=('Arial bold', 10))
        lblhd.grid(column=1, row=4)

        lbltkh = tk.Label(window, text='Ten khach hang', bg='gray', fg='white', font=('Arial', 10))
        lbltkh.grid(column=1, row=2)

        lblns = tk.Label(window, text='Nam sinh', bg='gray', fg='white', font=('Arial', 10))
        lblns.grid(column=5, row=2)

        lblqq = tk.Label(window, text='Que quan ', bg='gray', fg='white', font=('Arial', 10))
        lblqq.grid(column=7, row=2)

        lblcccd = tk.Label(window, text='So CCCD', bg='gray', fg='white', font=('Arial', 10))
        lblcccd.grid(column=1, row=3)

        lblgt = tk.Label(window, text='Gioi tinh', bg='gray', fg='white', font=('Arial', 10))
        lblgt.grid(column=3, row=2)

        lbldc = tk.Label(window, text='Dia chi', bg='gray', fg='white', font=('Arial', 10))
        lbldc.grid(column=5, row=3)

        lblnn = tk.Label(window, text='Nghe nghiep', bg='gray', fg='white', font=('Arial', 10))
        lblnn.grid(column=3, row=3)

        lblsdt = tk.Label(window, text='So dien thoai', bg='gray', fg='white', font=('Arial', 10))
        lblsdt.grid(column=7, row=3)

        lbllhd = tk.Label(window, text='Loai hop dong', bg='gray', fg='white', font=('Arial', 10))
        lbllhd.grid(column=1, row=5)

        lbllp = tk.Label(window, text='Loai phong', bg='gray', fg='white', font=('Arial', 10))
        lbllp.grid(column=3, row=5)

        lblps = tk.Label(window, text='So phong', bg='gray', fg='white', font=('Arial', 10))
        lblps.grid(column=5, row=5)

        lbltdc = tk.Label(window, text='Tien dat coc', bg='gray', fg='white', font=('Arial', 10))
        lbltdc.grid(column=1, row=6)

        lblnnp = tk.Label(window, text='Ngay nhan phong', bg='gray', fg='white', font=('Arial', 10))
        lblnnp.grid(column=3, row=6)

        lblntp = tk.Label(window, text='Ngay tra phong', bg='gray', fg='white', font=('Arial', 10))
        lblntp.grid(column=5, row=6)

        lblgp = tk.Label(window, text='Gia phong', bg='gray', fg='white', font=('Arial', 10))
        lblgp.grid(column=7, row=5)

        lbltt = tk.Label(window, text='Thanh toan', bg='gray', fg='white', font=('Arial', 10))
        lbltt.grid(column=7, row=6)

        lbdct= tk.LabelFrame(window, text='22 Hoang Van Hoe',bg='gray',fg='White',font=('Time new roman bold',20))
        lbdct.place(height=40, width=250,rely=0.85, relx=0.01)

        lbdct = tk.LabelFrame(window, text='THANHTUAN APARTMENT', bg='#116562', fg='White', font=('Time new roman bold', 30))
        lbdct.place(height=50, width=530, rely=0.78,relx=0.45)

        btin= tk.Button(window,text='In du lieu',bg='Blue', fg='white', font=('Arial bold', 10))
        btin.place(rely=0.88, relx=0.8)

        txttkh = tk.Entry(window, width=20)
        txttkh.grid(column=2, row=2)

        txtnn = tk.Entry(window, width=20)
        txtnn.grid(column=4, row=3)

        txtsdt = tk.Entry(window, width=20)
        txtsdt.grid(column=8, row=3)

        txtdc = tk.Entry(window, width=20)
        txtdc.grid(column=6, row=3)

        txttdc = tk.Entry(window, width=20)
        txttdc.grid(column=2, row=6)

        txtcccd = tk.Entry(window, width=20)
        txtcccd.grid(column=2, row=3)

        txtgp = tk.Entry(window, width=20)
        txtgp.grid(column=8, row=5)

        btn = tk.Button(window, text='Luu thong tin' ,bg='Blue', fg='white', font=('Arial bold', 10), command=save)
        btn.grid(column=7, row=7)

        btn1 = tk.Button(window, text='Xoa', bg='red', fg='white', font=('Arial bold',10), command=xnclear)
        btn1.grid(column=8, row=7)

        btn2 = tk.Button(window, text='Thoat', font=('Arial bold',10),bg='red',fg='white', command=quit)
        btn2.place(rely=0.88,relx=0.9)

        combolhd = ttk.Combobox(window, width=20)
        combolhd['values'] = ('Thue ngay', 'Thue thang', 'Thue nam')
        combolhd.grid(column=2, row=5)

        combogt = ttk.Combobox(window, width=20)
        combogt['values'] = ('Nam', 'Nu', 'Khac')
        combogt.grid(column=4, row=2)

        comboqq = ttk.Combobox(window, width=20)
        comboqq['values'] = ('An Giang', 'Bà Rịa-Vũng Tàu', 'Bạc Liêu', 'Bắc Kạn', 'Bắc Giang', 'Bắc Ninh', 'Bến Tre',
                             'Bình Dương', 'Bình Định', 'Bình Phước', 'Bình Thuận', 'Cà Mau', 'Cao Bằng',
                             'Cần Thơ (TP)',
                             'Đà Nẵng', '(TP)Đắk Lắk', 'Đắk Nông', 'Điện Biên', 'Đồng Nai', 'Đồng Tháp', 'Gia Lai',
                             'Hà Giang', 'Hà Nam', 'Hà Nội (TP)', 'Hà Tây',
                             'Hà Tĩnh', 'Hải Dương', 'Hải Phòng', '(TP)Hòa Bình', 'Hồ Chí Minh (TP)', 'Hậu Giang',
                             'Hưng Yên', 'Khánh Hòa', 'Kiên Giang', 'Kon Tum', 'Lai Châu', 'Lào Cai', 'Lạng Sơn',
                             'Lâm Đồng',
                             'Long An', 'Nam Định', 'Nghệ An', 'Ninh Bình', 'Ninh Thuận', 'Phú Thọ', 'Phú Yên',
                             'Quảng Bình', 'Quảng Nam', 'Quảng Ngãi', 'Quảng Ninh', 'Quảng Trị',
                             'Sóc Trăng', 'Sơn La', 'Tây Ninh', 'Thái Bình', 'Thái Nguyên', 'Thanh Hóa',
                             'Thừa Thiên – Huế', 'Tiền Giang', 'Trà Vinh', 'Tuyên Quang', 'Vĩnh Long', 'Vĩnh Phúc',
                             'Yên Bái')
        comboqq.grid(column=8, row=2)

        combolp = ttk.Combobox(window, width=20)
        combolp['values'] = ('Phong don', 'Phong doi', 'Phong da', 'Phong giuong tang')
        combolp.grid(column=4, row=5)

        combops = ttk.Combobox(window, width=20)
        combops['values'] = ('1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13',
                             '14', '15', '16', '17', '18', '19', '20')
        combops.grid(column=6, row=5)

        combott = ttk.Combobox(window, width=20)
        combott['values'] = ('Da thanh toan', 'Chua thanh toan')
        combott.grid(column=8, row=6)

        calNs = DateEntry(window, width=20)
        calNs.grid(column=6, row=2)

        calnnp = DateEntry(window, width=20)
        calnnp.grid(column=4, row=6)

        calntp = DateEntry(window, width=20)
        calntp.grid(column=6, row=6)



        frame1 = tk.LabelFrame(window, text='Du Lieu Excel', fg='Blue', font=('Arial bold', 10))
        frame1.place(height=400, width=1000, rely=0.2, relx=0.01)

        file_frame = tk.LabelFrame(window, text="Mo File", fg='Blue', font=('Arial bold', 10))
        file_frame.place(height=100, width=400, rely=0.7, relx=0.01)

        button1 = tk.Button(file_frame, text="Chon mot file", bg='gray', fg='white', font=('Arial', 10), command=lambda: File_dialog())
        button1.place(rely=0.65, relx=0.75)

        button2 = tk.Button(file_frame, text="Tai du lieu", bg='gray', fg='white', font=('Arial', 10), command=lambda: Load_excel_data())
        button2.place(rely=0.65, relx=0.55)

        label_file = ttk.Label(file_frame, text="Chua co file nao duoc chon",font=('Arial',10))
        label_file.place(rely=0, relx=0)

        tv1 = ttk.Treeview(frame1)
        tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(frame1, orient="vertical",
                                   command=tv1.yview)
        treescrollx = tk.Scrollbar(frame1, orient="horizontal",
                                   command=tv1.xview)
        tv1.configure(xscrollcommand=treescrollx.set,
                      yscrollcommand=treescrolly.set)
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")

        def File_dialog():
            filename = filedialog.askopenfilename(initialdir="/",
                                                  title="Select A File",
                                                  filetype=(("xlsx files", "*.xlsx"), ("All Files", "*.*")))
            label_file["text"] = filename
            return None

        def Load_excel_data():
            file_path = label_file["text"]
            try:
                excel_filename = r"{}".format(file_path)
                if excel_filename[-4:] == ".csv":
                    df = pd.read_csv(excel_filename)
                else:
                    df = pd.read_excel(excel_filename)

            except ValueError:
                tk.messagebox.showerror("Thong bao!", "File ban chon khong hop le")
                return None
            except FileNotFoundError:
                tk.messagebox.showerror("Thong bao!", f"Khong co file nhu vay {file_path}")
                return None

            clear_data()
            tv1["column"] = list(df.columns)
            tv1["show"] = "headings"
            for column in tv1["columns"]:
                tv1.heading(column, text=column)

            df_rows = df.to_numpy().tolist()
            for row in df_rows:
                tv1.insert("", "end",
                           values=row)
            return None

        def clear_data():
            tv1.delete(*tv1.get_children())
            return None




        window.mainloop()


    else:
        messagebox.showinfo('That bai!', ' Vui long dang nhap lai')


global entry1
global entry2

lb = tk.Label(window, text='Tai khoan',bg='orange', fg='white', font=('Arial bold', 10),bd=5)
lb.place(x=30, y=20)

lb1 = tk.Label(window, text='Mat khau',bg='orange', fg='white', font=('Arial bold', 10),bd=5)
lb1.place(x=30, y=70)

entry1 = tk.Entry(window, bd=5)
entry1.place(x=140, y=20)
entry2 = tk.Entry(window, bd=5)
entry2.place(x=140, y=70)

bt = tk.Button(window, text='Dang nhap', bg='brown',fg='white',font=('Arial bold',10), command=login, height=2, width=10, bd=6)
bt.place(x=100, y=120)


window.mainloop()
